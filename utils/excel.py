# utils/excel.py — Generación de archivos Excel (.xlsx) para reportes COCOMO
# Extraído de app.py — funciones puras de openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime


# ═══════════════════════════════════════════
#   ESTILOS
# ═══════════════════════════════════════════

ESTILOS = {
    "titulo":       Font(name="Calibri", size=16, bold=True, color="1F4E79"),
    "subtitulo":    Font(name="Calibri", size=11, color="5A5F70"),
    "cabecera":     Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
    "celda":        Font(name="Calibri", size=11),
    "celda_num":    Font(name="Consolas", size=11),
    "fill_cab":     PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid"),
    "fill_resalt":  PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
    "fill_claro":   PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid"),
    "borde":        Border(left=Side(style="thin", color="C0C0C0"),
                           right=Side(style="thin", color="C0C0C0"),
                           top=Side(style="thin", color="C0C0C0"),
                           bottom=Side(style="thin", color="C0C0C0")),
}


# ═══════════════════════════════════════════
#   HELPERS DE CELDAS
# ═══════════════════════════════════════════

def _encabezados(ws, row, cols):
    """Escribe una fila de encabezados con estilo."""
    for c, texto in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=texto)
        cell.font = ESTILOS["cabecera"]
        cell.fill = ESTILOS["fill_cab"]
        cell.alignment = Alignment(horizontal="center")
        cell.border = ESTILOS["borde"]


def _fila(ws, row, datos, num_cols=None, destacar=False):
    """Escribe una fila de datos con bordes."""
    if num_cols is None:
        num_cols = set()
    for c, val in enumerate(datos, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = ESTILOS["celda_num"] if c in num_cols else ESTILOS["celda"]
        cell.border = ESTILOS["borde"]
        cell.alignment = Alignment(horizontal="center")
        if destacar:
            cell.fill = ESTILOS["fill_resalt"]


def _titulo(ws, row, texto, col_fin=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_fin)
    cell = ws.cell(row=row, column=1, value=texto)
    cell.font = ESTILOS["titulo"]


def _subtitulo(ws, row, texto, col_fin=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_fin)
    cell = ws.cell(row=row, column=1, value=texto)
    cell.font = ESTILOS["subtitulo"]


def _ajustar(ws, cols_widths):
    for c, w in enumerate(cols_widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w


def _guardar_en_bytesio(wb):
    """Guarda el workbook en un BytesIO y lo rebobina para send_file."""
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════
#   BUILDERS DE EXCEL POR MODELO
# ═══════════════════════════════════════════

def build_c81basico(data):
    """Genera workbook Excel para COCOMO 81 Básico."""
    body = data["body"]
    r = data["resultado"]
    p = r["params"]
    etapas = data["etapas"]

    wb = Workbook()
    ws = wb.active
    ws.title = "COCOMO 81 Basico"
    nc = 5

    _titulo(ws, 1, "COCOMO 81 — Modelo Basico", nc)
    _subtitulo(ws, 2, f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Modo: {body['modo']}  |  EAF = 1.0 (fijo)", nc)

    row = 4
    _encabezados(ws, row, ["Parametros de Entrada", "", "", "", ""])
    row += 1
    _fila(ws, row, ["Tamano (KLOC)", body["kloc"], "", "", ""])
    row += 1
    _fila(ws, row, ["Modo del proyecto", body["modo"].capitalize(), "", "", ""])
    row += 1
    _fila(ws, row, ["Costo hombre-mes ($)", f"${body['costo_hm']:,.2f}", "", "", ""])
    row += 1
    _fila(ws, row, ["Constantes", f"a={p['a']}  b={p['b']}  c={p['c']}  d={p['d']}", "", "", ""])

    row += 2
    _encabezados(ws, row, ["Resultados", "Valor", "Unidad", "", ""])
    row += 1
    _fila(ws, row, ["Esfuerzo", round(r["esfuerzo"], 2), "persona-mes (PM)", "", ""], num_cols={2}, destacar=True)
    row += 1
    _fila(ws, row, ["Tiempo de desarrollo", round(r["tiempo"], 2), "meses", "", ""], num_cols={2})
    row += 1
    _fila(ws, row, ["Personal requerido", round(r["personal"], 2), "personas", "", ""], num_cols={2})
    row += 1
    _fila(ws, row, ["Costo total", f"${r['costo']:,.2f}", "USD", "", ""], num_cols={2}, destacar=True)

    row += 2
    _encabezados(ws, row, ["Distribucion por Etapas", "Esfuerzo (PM)", "Tiempo (meses)", "% Esfuerzo", ""])
    for e in etapas:
        row += 1
        _fila(ws, row, [e["nombre"], e["esfuerzo"], e["tiempo"], f"{e['pct']}%", ""], num_cols={2, 3})

    _ajustar(ws, [36, 22, 22, 16, 16])

    return _guardar_en_bytesio(wb), f"COCOMO81_Basico_{datetime.now().strftime('%Y%m%d')}.xlsx"


def build_c81(data):
    """Genera workbook Excel para COCOMO 81 Intermedio."""
    body = data["body"]
    r = data["resultado"]
    p = r["params"]
    etapas = data["etapas"]
    eaf = data["eaf"]
    drivers = data.get("valoresDrivers", {})

    wb = Workbook()
    ws = wb.active
    ws.title = "COCOMO 81 Intermedio"
    nc = 5

    _titulo(ws, 1, "COCOMO 81 — Modelo Intermedio", nc)
    _subtitulo(ws, 2, f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Modo: {body['modo']}  |  EAF = {eaf:.4f}", nc)

    row = 4
    _encabezados(ws, row, ["Parametros de Entrada", "", "", "", ""])
    row += 1
    _fila(ws, row, ["Tamano (KLOC)", body["kloc"], "", "", ""])
    row += 1
    _fila(ws, row, ["Modo del proyecto", body["modo"].capitalize(), "", "", ""])
    row += 1
    _fila(ws, row, ["Costo hombre-mes ($)", f"${body['costo_hm']:,.2f}", "", "", ""])
    row += 1
    _fila(ws, row, ["Constantes", f"a={p['a']}  b={p['b']}  c={p['c']}  d={p['d']}", "", "", ""])

    row += 2
    _encabezados(ws, row, ["Conductor de Coste", "Codigo", "Valor", "", ""])
    for code, info in drivers.items():
        row += 1
        _fila(ws, row, [info.get("nombre", code), code, info["valor"], "", ""], num_cols={3})
    row += 1
    _fila(ws, row, ["EAF TOTAL (producto)", "", round(eaf, 4), "", ""], num_cols={3}, destacar=True)

    row += 2
    _encabezados(ws, row, ["Resultados", "Valor", "Unidad", "", ""])
    row += 1
    _fila(ws, row, ["Esfuerzo", round(r["esfuerzo"], 2), "persona-mes (PM)", "", ""], num_cols={2}, destacar=True)
    row += 1
    _fila(ws, row, ["Tiempo de desarrollo", round(r["tiempo"], 2), "meses", "", ""], num_cols={2})
    row += 1
    _fila(ws, row, ["Personal requerido", round(r["personal"], 2), "personas", "", ""], num_cols={2})
    row += 1
    _fila(ws, row, ["Costo total", f"${r['costo']:,.2f}", "USD", "", ""], num_cols={2}, destacar=True)

    row += 2
    _encabezados(ws, row, ["Distribucion por Etapas", "Esfuerzo (PM)", "Tiempo (meses)", "% Esfuerzo", ""])
    for e in etapas:
        row += 1
        _fila(ws, row, [e["nombre"], e["esfuerzo"], e["tiempo"], f"{e['pct']}%", ""], num_cols={2, 3})

    _ajustar(ws, [36, 22, 22, 16, 16])

    return _guardar_en_bytesio(wb), f"COCOMO81_Intermedio_{datetime.now().strftime('%Y%m%d')}.xlsx"


def build_cocomo2(data):
    """Genera workbook Excel para COCOMO II."""
    body = data["body"]
    r = data["resultado"]
    etapas = data["etapas"]
    B_val = data["B"]
    EM_val = data["EM"]
    sumSF = data["sumSF"]

    wb = Workbook()
    ws = wb.active
    ws.title = "COCOMO II"
    nc = 5

    _titulo(ws, 1, "COCOMO II — Post-Arquitectura", nc)
    _subtitulo(ws, 2, f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  B = {B_val:.4f}  |  EM = {EM_val:.4f}", nc)

    row = 4
    _encabezados(ws, row, ["Parametros de Entrada", "", "", "", ""])
    row += 1
    _fila(ws, row, ["Tamano (KLOC)", body["kloc"], "", "", ""])
    row += 1
    _fila(ws, row, ["Costo hombre-mes ($)", f"${body['costo_hm']:,.2f}", "", "", ""])
    row += 1
    _fila(ws, row, ["Suma Factores de Escala (SF)", sumSF, "", "", ""])
    row += 1
    _fila(ws, row, ["Exponente B = 0.91 + 0.01 x SF", round(B_val, 4), "", "", ""])
    row += 1
    _fila(ws, row, ["Producto EM", round(EM_val, 4), "", "", ""])

    row += 2
    sf_names = ["PREC - Precedencia", "FLEX - Flexibilidad", "RESL - Arquitectura/Resolucion", "TEAM - Cohesion equipo", "PMAT - Madurez proceso"]
    _encabezados(ws, row, ["Factor de Escala (SF)", "Valor", "", "", ""])
    for name, val in zip(sf_names, body["sf_valores"]):
        row += 1
        _fila(ws, row, [name, val, "", "", ""], num_cols={2})

    row += 2
    em_names = [
        "RELY - Confiabilidad", "DATA - Tamano BD", "CPLX - Complejidad", "RUSE - Reutilizacion",
        "DOCU - Documentacion", "TIME - Restriccion CPU", "STOR - Restriccion almacenamiento",
        "PVOL - Volatilidad plataforma", "ACAP - Capacidad analistas", "PCAP - Capacidad programadores",
        "PCON - Continuidad personal", "APEX - Experiencia aplicacion", "PLEX - Experiencia plataforma",
        "LTEX - Experiencia lenguaje", "TOOL - Uso herramientas", "SITE - Multisitio", "SCED - Compresion calendario"
    ]
    _encabezados(ws, row, ["Multiplicador de Esfuerzo (EM)", "Valor", "", "", ""])
    for name, val in zip(em_names, body["em_valores"]):
        row += 1
        _fila(ws, row, [name, val, "", "", ""], num_cols={2})

    row += 2
    _encabezados(ws, row, ["Resultados", "Valor", "Unidad", "", ""])
    row += 1
    _fila(ws, row, ["Esfuerzo", round(r["esfuerzo"], 2), "persona-mes (PM)", "", ""], num_cols={2}, destacar=True)
    row += 1
    _fila(ws, row, ["Tiempo de desarrollo (TDEV)", round(r["tiempo"], 2), "meses", "", ""], num_cols={2})
    row += 1
    _fila(ws, row, ["Personal requerido", round(r["personal"], 2), "personas", "", ""], num_cols={2})
    row += 1
    _fila(ws, row, ["Costo total", f"${r['costo']:,.2f}", "USD", "", ""], num_cols={2}, destacar=True)

    row += 2
    _encabezados(ws, row, ["Distribucion por Etapas", "Esfuerzo (PM)", "Tiempo (meses)", "% Esfuerzo", ""])
    for e in etapas:
        row += 1
        _fila(ws, row, [e["nombre"], e["esfuerzo"], e["tiempo"], f"{e['pct']}%", ""], num_cols={2, 3})

    _ajustar(ws, [40, 22, 22, 16, 16])

    return _guardar_en_bytesio(wb), f"COCOMO_II_{datetime.now().strftime('%Y%m%d')}.xlsx"


def build_fp(data):
    """Genera workbook Excel para Puntos de Funcion."""
    from models.puntos_funcion import VAF_CARACTERISTICAS

    body = data["body"]
    r = data["resultado"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Puntos de Funcion"
    nc = 5

    _titulo(ws, 1, "Puntos de Funcion (FPA)", nc)
    _subtitulo(ws, 2, f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}", nc)

    row = 4
    _encabezados(ws, row, ["Componente", "Simple", "Medio", "Complejo", ""])
    fp_types = ["EI - Entradas externas", "EO - Salidas externas", "EQ - Consultas externas",
                "ILF - Archivos logicos int.", "EIF - Archivos interfaz ext."]
    for name, (s, m, c) in zip(fp_types, body["conteos"].values()):
        row += 1
        _fila(ws, row, [name, s, m, c, ""], num_cols={2, 3, 4})

    row += 2
    _encabezados(ws, row, ["Caracteristica VAF", "Valor (0-5)", "", "", ""])
    for i, val in enumerate(body["vaf_valores"]):
        row += 1
        _fila(ws, row, [VAF_CARACTERISTICAS[i], val, "", "", ""], num_cols={2})

    row += 2
    _encabezados(ws, row, ["Resultados", "Valor", "Unidad", "", ""])
    row += 1
    _fila(ws, row, ["PF Sin Ajustar (PFSA)", r["pfsa"], "puntos funcion", "", ""], num_cols={2}, destacar=True)
    row += 1
    _fila(ws, row, ["Factor de Ajuste (VAF)", r["vaf"], "", "", ""], num_cols={2})
    row += 1
    _fila(ws, row, ["PF Ajustados (PFA)", r["pfa"], "puntos funcion", "", ""], num_cols={2}, destacar=True)
    row += 1
    _fila(ws, row, ["KLOC estimadas", r["kloc"], "miles de lineas", "", ""], num_cols={2}, destacar=True)

    _ajustar(ws, [36, 20, 22, 16, 16])

    return _guardar_en_bytesio(wb), f"Puntos_Funcion_{datetime.now().strftime('%Y%m%d')}.xlsx"
